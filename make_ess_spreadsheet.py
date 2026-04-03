"""Build ESS_explained.xlsx — run once; open in Excel."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "ESS explained"

header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")
note_font = Font(italic=True)

def hdr(r, c, text):
    cell = ws.cell(r, c, text)
    cell.fill = header_fill
    cell.font = header_font

# --- Title / explanation ---
ws.merge_cells("A1:F1")
ws["A1"] = "Effective Sample Size (ESS) — very simple"
ws["A1"].font = Font(bold=True, size=14)

ws.merge_cells("A2:F5")
ws["A2"] = (
    "Each Monte Carlo sample k has a complex weight W_k = (real part) + i×(imag part).\n"
    "ESS/N measures how well those weights combine (like arrows).\n\n"
    "Formula (same idea as your notebook):  ESS/N = |average(W)|² / average(|W|²)\n"
    "• If all W point the same way and have similar size → ESS/N near 1 (great).\n"
    "• If they cancel or a few dominate → ESS/N small (bad).\n"
    "|W|² = (real)² + (imag)² for each row."
)
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

# --- Table: good case (aligned weights) ---
start = 7
ws.cell(start, 1, "Example A — weights similar (good thimble)")
ws.cell(start, 1).font = Font(bold=True)
start += 1

hdr(start, 1, "k")
hdr(start, 2, "W real")
hdr(start, 3, "W imag")
hdr(start, 4, "|W|²")
start += 1

# Rows 8-12: data for good case
good_data = [
    (1, 1.0, 0.1),
    (2, 0.95, 0.12),
    (3, 1.02, 0.08),
    (4, 0.98, 0.11),
    (5, 1.01, 0.09),
]
row0 = start
for k, re, im in good_data:
    ws.cell(start, 1, k)
    ws.cell(start, 2, re)
    ws.cell(start, 3, im)
    ws.cell(start, 4, f"=B{start}*B{start}+C{start}*C{start}")
    start += 1
end_data = start - 1

ws.cell(start, 1, "average(W) real")
ws.cell(start, 2, f"=AVERAGE(B{row0}:B{end_data})")
ws.cell(start, 3, f"=AVERAGE(C{row0}:C{end_data})")
ws.cell(start, 1).font = Font(bold=True)
avg_re_row = start
start += 1

ws.cell(start, 1, "|average(W)|²")
ws.cell(start, 2, f"=B{avg_re_row}*B{avg_re_row}+C{avg_re_row}*C{avg_re_row}")
ws.cell(start, 1).font = Font(bold=True)
mod_mean_row = start
start += 1

ws.cell(start, 1, "average(|W|²)")
ws.cell(start, 2, f"=AVERAGE(D{row0}:D{end_data})")
ws.cell(start, 1).font = Font(bold=True)
avg_modsq_row = start
start += 1

ws.cell(start, 1, "ESS / N")
ws.cell(start, 2, f"=B{mod_mean_row}/B{avg_modsq_row}")
ws.cell(start, 1).font = Font(bold=True)
ws.cell(start, 2).number_format = "0.0000"

# --- Table: bad case (one huge weight) ---
start += 2
ws.cell(start, 1, "Example B — one weight huge (bad sampling)")
ws.cell(start, 1).font = Font(bold=True)
start += 1

hdr(start, 1, "k")
hdr(start, 2, "W real")
hdr(start, 3, "W imag")
hdr(start, 4, "|W|²")
start += 1

bad_data = [
    (1, 0.1, 0.01),
    (2, 0.11, 0.01),
    (3, 0.09, 0.02),
    (4, 0.1, 0.01),
    (5, 9.5, 0.5),  # dominates
]
row0_b = start
for k, re, im in bad_data:
    ws.cell(start, 1, k)
    ws.cell(start, 2, re)
    ws.cell(start, 3, im)
    ws.cell(start, 4, f"=B{start}*B{start}+C{start}*C{start}")
    start += 1
end_data_b = start - 1

ws.cell(start, 1, "average(W) real")
ws.cell(start, 2, f"=AVERAGE(B{row0_b}:B{end_data_b})")
ws.cell(start, 3, f"=AVERAGE(C{row0_b}:C{end_data_b})")
ws.cell(start, 1).font = Font(bold=True)
avg_re_b = start
start += 1

ws.cell(start, 1, "|average(W)|²")
ws.cell(start, 2, f"=B{avg_re_b}*B{avg_re_b}+C{avg_re_b}*C{avg_re_b}")
ws.cell(start, 1).font = Font(bold=True)
mod_mean_b = start
start += 1

ws.cell(start, 1, "average(|W|²)")
ws.cell(start, 2, f"=AVERAGE(D{row0_b}:D{end_data_b})")
ws.cell(start, 1).font = Font(bold=True)
avg_modsq_b = start
start += 1

ws.cell(start, 1, "ESS / N")
ws.cell(start, 2, f"=B{mod_mean_b}/B{avg_modsq_b}")
ws.cell(start, 1).font = Font(bold=True)
ws.cell(start, 2).number_format = "0.0000"

# column widths
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 18

out = r"c:\xampp1\Quantum\ESS_explained.xlsx"
wb.save(out)
print("Wrote", out)
